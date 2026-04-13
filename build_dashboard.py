import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.protection import SheetProtection
import os

LOGO_PATH = r"C:\Users\ameer\.verdent\artifacts\buckets\4bae9bb5-302f-4f09-9754-dbb0067de45e\images\1775991837070_eb8e8585.jpeg"
OUTPUT_PATH = r"C:\Users\ameer\.verdent\verdent-projects\new-project\PMO_Dashboard.xlsx"

TEAL = "0D7377"
DARK_TEAL = "0A5C5F"
LIGHT_TEAL = "E0F2F1"
ACCENT_TEAL = "26A69A"
WHITE = "FFFFFF"
DARK_GRAY = "333333"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "BDBDBD"
RED = "E53935"
GREEN = "43A047"
ORANGE = "FB8C00"
BLUE = "1565C0"

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)

header_font = Font(name='Calibri', bold=True, size=12, color=WHITE)
header_fill = PatternFill('solid', fgColor=DARK_TEAL)
subheader_font = Font(name='Calibri', bold=True, size=11, color=WHITE)
subheader_fill = PatternFill('solid', fgColor=TEAL)
data_font = Font(name='Calibri', size=11, color=DARK_GRAY)
title_font = Font(name='Calibri', bold=True, size=18, color=DARK_TEAL)
subtitle_font = Font(name='Calibri', bold=True, size=14, color=TEAL)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
wrap_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

EMPLOYEES = [
    {
        "name": "شكرية كنكار",
        "portfolios": ["المالية", "الحوكمة والمخاطر والالتزام وحوكمة البيانات", "المنافع"],
        "goals": [
            "إنجاز اكتمال واغلاق المشاريع للمحافظ التالية (المالية, الحوكمة والمخاطر والالتزام وحوكمة البيانات والمنافع) بنسبة 100% مع عكس البيانات الصحيحة للمشاريع في منصة ميسر ومشروعي حتى نهاية ديسمبر 2026",
            "تحقيق الالتزام بخطة الصرف المالية مع الالتزام بالجدول الزمني للمشاريع كما هو مخطط له بنسبة 100% للمشاريع المتعاقد عليها للمحافظ التالية (المالية, الحوكمة والمخاطر والالتزام وحوكمة البيانات, المنافع والنتائج الصحية) بحلول نهاية ديسمبر لعام 2026",
            "تحقيق الالتزام بإغلاق كافة التغييرات والمخاطر والتحديات بنسبة 100% للمشاريع القائمة والمعتمدة للمحافظ التالية (المالية, الحوكمة والمخاطر والالتزام وحوكمة البيانات, المنافع والنتائج الصحية) بحلول نهاية ديسمبر لعام 2026",
            "الالتزام بالتقارير الأسبوعية ومتابعة وتحديث المشاريع التشغيلية والاستراتيجية والمطروحة للمحافظ التالية (المالية, الحوكمة والمخاطر والالتزام وحوكمة البيانات, المنافع والنتائج الصحية) بعدد 40 تقرير لكل قطاع الى نهاية ديسمبر لعام 2026",
            "إقامة عدد ثلاث ورش عمل او تدريب تختص بإدارة المشاريع والانظمة المرتبطة بها للمحافظ التالية (المالية, الحوكمة والمخاطر والالتزام وحوكمة البيانات والمنافع), ومبادرة مرتبطة بإدارة المشاريع وكذلك أهداف القطاع والمركز بحلول نهاية ديسمبر لعام 2026"
        ],
        "projects": [
            ("تدقيق وإجازة عمليات الشراء من القطاع الخاص", "المالية"),
            ("بناء وتفعيل قطاع الشؤون المالية لمركز التأمين الصحي الوطني", "المالية"),
            ("مشروع تطوير ضوابط أمان التطبيقات والبنية التحتية", "الحوكمة"),
            ("حصر وجرد وترميز وتقييم الأصول لمركز التأمين الصحي الوطني", "المالية"),
            ("الاحتساب الاكتواري لرصيد مخصص مكافأة نهاية الخدمة لأغراض اعداد القوائم المالية", "المالية"),
            ("تنسيق تنفيذ عمليات إدارة صحة السكان بشكل تكاملي", "المنافع"),
            ("تطوير لائحة وثائق تغطية الرعاية الصحية", "القانونية"),
        ]
    },
    {
        "name": "نايف الجلال",
        "portfolios": ["الشراء الاستراتيجي", "الاستدامة المالية", "القانونية", "الاستراتيجية"],
        "goals": [
            "إنجاز اكتمال واغلاق المشاريع للمحافظ التالية (الاستدامة المالية, القانونية, الاستراتيجية, الشراء الاستراتيجي) بنسبة 100% مع عكس البيانات الصحيحة للمشاريع في منصة ميسر ومشروعي حتى نهاية ديسمبر 2026",
            "تحقيق الالتزام بخطة الصرف المالية مع الالتزام بالجدول الزمني للمشاريع كما هو مخطط له بنسبة 100% للمشاريع المتعاقد عليها للمحافظ التالية (الاستدامة المالية, القانونية والاستراتيجية) بحلول نهاية ديسمبر لعام 2026",
            "تحقيق الالتزام بإغلاق التغييرات والمخاطر والتحديات بنسبة 100% للمشاريع القائمة والمعتمدة للمحافظ التالية (الاستدامة المالية, القانونية, الاستراتيجية والشراء الاستراتيجي) بحلول نهاية ديسمبر لعام 2026",
            "الالتزام بالتقارير الأسبوعية ومتابعة وتحديث المشاريع التشغيلية والاستراتيجية والمطروحة للمحافظ التالية (الاستدامة المالية, القانونية والاستراتيجية, الشراء الاستراتيجي) بعدد 40 تقرير لكل قطاع الى نهاية ديسمبر لعام 2026",
            "إقامة عدد ثلاث ورش عمل او تدريب تختص بإدارة المشاريع والانظمة المرتبطة بها للمحافظ التالية (الاستدامة المالية, القانونية والاستراتيجية, الشراء الاستراتيجي), ومبادرة مرتبطة بإدارة المشاريع وكذلك أهداف القطاع والمركز بحلول نهاية ديسمبر لعام 2026"
        ],
        "projects": [
            ("تصميم وتشغيل استراتيجية مركز التأمين الصحي الوطني للشراء من القطاع الخاص", "الشراء الاستراتيجي"),
            ("مشروع دعم تطوير التصور الشامل حول التأمين على المواطنين", "الاستراتيجية"),
            ("إدارة عمليات شراء الخدمات الطبية من القطاع الخاص-الإحالة الداخلية ومراجعة المطالبات", "الشراء الاستراتيجي"),
            ("ادارة الاحالات والمطالبات الطبية للمرضى الذين يتم علاجهم داخل المملكة العربية السعودية", "الشراء الاستراتيجي"),
            ("ادارة الاحالات والمطالبات الطبية للمرضى الذين يتم علاجهم خارج المملكة العربية السعودية", "الشراء الاستراتيجي"),
            ("إنشاء وتفعيل وحدة الشفافية المالية والاستدامة التابعة لمركز التأمين الصحي الوطني", "الاستدامة المالية"),
        ]
    },
    {
        "name": "عبد العزيز المعجل",
        "portfolios": ["المراجعة الداخلية", "تقييم التقنيات الصحية", "مكتب الرئيس التنفيذي"],
        "goals": [
            "إنجاز اكتمال واغلاق المشاريع للمحافظ التالية (المراجعة الداخلية, مركز تقييم التقنيات الصحية ومكتب الرئيس التنفيذي) بنسبة 100% مع عكس البيانات الصحيحة للمشاريع في منصة ميسر ومشروعي حتى نهاية ديسمبر 2026",
            "تحقيق الالتزام بخطة الصرف المالية مع الالتزام بالجدول الزمني للمشاريع كما هو مخطط له بنسبة 100% للمشاريع المتعاقد عليها للمحافظ التالية (المراجعة الداخلية, مركز تقييم التقنيات الصحية ومكتب الرئيس التنفيذي) بحلول نهاية ديسمبر لعام 2026",
            "تحقيق الالتزام بإغلاق التغييرات والمخاطر والتحديات بنسبة 100% للمشاريع القائمة والمعتمدة للمحافظ التالية (المراجعة الداخلية, مركز تقييم التقنيات الصحية ومكتب الرئيس التنفيذي) بحلول نهاية ديسمبر لعام 2026",
            "الالتزام بالتقارير الأسبوعية ومتابعة وتحديث المشاريع التشغيلية والاستراتيجية والمطروحة للمحافظ التالية (المراجعة الداخلية, مركز تقييم التقنيات الصحية ومكتب الرئيس التنفيذي) بعدد 40 تقرير لكل قطاع الى نهاية ديسمبر لعام 2026",
            "إقامة عدد ثلاث ورش عمل او تدريب تختص بإدارة المشاريع والانظمة المرتبطة بها للمحافظ التالية (المراجعة الداخلية, مركز تقييم التقنيات الصحية ومكتب الرئيس التنفيذي), ومبادرة مرتبطة بإدارة المشاريع وكذلك أهداف القطاع والمركز بحلول نهاية ديسمبر لعام 2026"
        ],
        "projects": [
            ("مشروع انشاء وحدة التحكم والتحول في مركز التأمين الصحي الوطني", "مكتب الرئيس التنفيذي"),
            ("إعادة تأسيس وتحديث إدارة المراجعة الداخلية في مركز التأمين الصحي", "المراجعة الداخلية"),
            ("دعم منهجية مركز تقييم التقنيات الصحية", "تقييم التقنيات الصحية"),
        ]
    },
    {
        "name": "صالح الغفيلي",
        "portfolios": ["الصحة الحكيمة", "الترميز الطبي وحساب التكاليف", "التمويل والتكاليف", "التقنية والذكاء الاصطناعي"],
        "goals": [
            "إنجاز اكتمال واغلاق المشاريع للمحافظ التالية (الصحة الحكيمة, الترميز الطبي وحساب التكاليف, التمويل والتكاليف, التقنية والذكاء الاصطناعي) بنسبة 100% مع عكس البيانات الصحيحة للمشاريع في منصة ميسر ومشروعي حتى نهاية ديسمبر 2026",
            "تحقيق الالتزام بخطة الصرف المالية مع الالتزام بالجدول الزمني للمشاريع كما هو مخطط له بنسبة 100% للمشاريع المتعاقد عليها للمحافظ التالية (الصحة الحكيمة, الترميز الطبي وحساب التكاليف, التمويل والتكاليف, التقنية والذكاء الاصطناعي) بحلول نهاية ديسمبر لعام 2026",
            "تحقيق الالتزام بإغلاق التغييرات والمخاطر والتحديات بنسبة 100% للمشاريع القائمة والمعتمدة للمحافظ التالية (الصحة الحكيمة, الترميز الطبي وحساب التكاليف, التمويل والتكاليف, التقنية والذكاء الاصطناعي) بحلول نهاية ديسمبر لعام 2026",
            "الالتزام بالتقارير الأسبوعية ومتابعة وتحديث المشاريع التشغيلية والاستراتيجية والمطروحة للمحافظ التالية (الصحة الحكيمة, الترميز الطبي وحساب التكاليف, التمويل والتكاليف, التقنية والذكاء الاصطناعي) بعدد 40 تقرير لكل قطاع الى نهاية ديسمبر لعام 2026",
            "إقامة عدد ثلاث ورش عمل او تدريب تختص بإدارة المشاريع والانظمة المرتبطة بها للمحافظ التالية (الصحة الحكيمة, الترميز الطبي وحساب التكاليف, التمويل والتكاليف, التقنية والذكاء الاصطناعي), ومبادرة مرتبطة بإدارة المشاريع وكذلك أهداف القطاع والمركز بحلول نهاية ديسمبر لعام 2026"
        ],
        "projects": [
            ("مشروع تفعيل دور مركز التميز الوطني للترميز الطبي وحساب التكاليف", "الترميز الطبي وحساب التكاليف"),
            ("تفعيل وبناء نماذج الدفع وشراء الخدمات الصحية المرحلة الأولى", "التقنية والذكاء الاصطناعي"),
            ("تطوير الموقع الخارجي وتقديم الدعم الفني", "التقنية والذكاء الاصطناعي"),
            ("منصة تكامل البيانات ومنصة عمل البيانات", "التقنية والذكاء الاصطناعي"),
            ("منصة التنبيهات الموحدة", "التقنية والذكاء الاصطناعي"),
            ("أنظمة خدمات تقنية المعلومات", "التقنية والذكاء الاصطناعي"),
            ("المرحلة الثانية من نظام التوظيف الإلكتروني", "التقنية والذكاء الاصطناعي"),
            ("استضافة البنية التحتية لمركز التأمين الصحي الوطني", "التقنية والذكاء الاصطناعي"),
            ("اعداد نظام التلفونات لمركز التأمين الصحي الوطني", "التقنية والذكاء الاصطناعي"),
            ("مراجعة وتحديث معايير ترميز الأمراض والإجراءات الطبية", "الترميز الطبي وحساب التكاليف"),
            ("مشروع البوابة الإلكترونية لجمع البيانات الطبية المرمزة (المرحلة الثالثة)", "التقنية والذكاء الاصطناعي"),
            ("مشروع تأسيس البنية التحتية لمركز بيانات برنامج الضمان الصحي وشراء الخدمات الصحية", "التقنية والذكاء الاصطناعي"),
            ("مشروع نظام الاتصالات الإدارية وإدارة اللجان والاجتماعات", "التقنية والذكاء الاصطناعي"),
            ("مشروع بناء قدرات تحليل البيانات وتشغيلها لبرنامج الضمان الصحي وشراء الخدمات", "التقنية والذكاء الاصطناعي"),
        ]
    },
    {
        "name": "مشعل محمد",
        "portfolios": ["الموارد البشرية", "التواصل والعناية بالمستفيدين"],
        "goals": [
            "إنجاز اكتمال واغلاق المشاريع للمحافظ التالية (الموارد البشرية, التواصل والعناية بالمستفيدين) بنسبة 100% مع عكس البيانات الصحيحة للمشاريع في منصة ميسر ومشروعي حتى نهاية ديسمبر 2026",
            "تحقيق الالتزام بخطة الصرف المالية والالتزام بالجدول الزمني للمشاريع كما هو مخطط له بنسبة 100% للمشاريع المتعاقد عليها للمحافظ التالية (الموارد البشرية, التواصل والعناية بالمستفيدين) بحلول نهاية ديسمبر لعام 2026",
            "تحقيق الالتزام بإغلاق التغييرات والمخاطر والتحديات بنسبة 100% للمشاريع القائمة والمعتمدة للمحافظ التالية (الموارد البشرية, التواصل والعناية بالمستفيدين) بحلول نهاية ديسمبر لعام 2026",
            "الالتزام بالتقارير الأسبوعية ومتابعة وتحديث المشاريع التشغيلية والاستراتيجية والمطروحة للمحافظ التالية (الموارد البشرية, التواصل والعناية بالمستفيدين) بعدد 40 تقرير لكل قطاع الى نهاية ديسمبر لعام 2026",
            "إقامة عدد ثلاث ورش عمل او تدريب تختص بإدارة المشاريع والانظمة المرتبطة بها للمحافظ التالية (الموارد البشرية, التواصل والعناية بالمستفيدين), ومبادرة مرتبطة بإدارة المشاريع وكذلك أهداف القطاع والمركز بحلول نهاية ديسمبر لعام 2026"
        ],
        "projects": [
            ("خدمات الصيانة والتشغيل لمركز التأمين الصحي الوطني", "الموارد البشرية"),
            ("تقديم الخدمات الاستشارية والدعم المساند لعمليات التشغيل لمركز التأمين الصحي الوطني", "الموارد البشرية"),
            ("مشروع تفعيل أدوار مركز التأمين الصحي الوطني في المنظومة الصحية المرحلة الأولى", "الموارد البشرية"),
            ("الحوكمة ونموذج التشغيل والمخطط الرقمي", "التواصل والعناية بالمستفيدين"),
        ]
    },
]

WEEKLY_TASKS = [
    ("هل تم تحديث الأنظمة والمنصات المتعلقة بالمشاريع؟", 1),
    ("هل تم تحديث الخطة التفصيلية للمشروع؟", 1),
    ("هل تم حل التحديات والمخاطر؟", 3),
    ("هل تم متابعة وإغلاق أوامر التغيير؟", 3),
    ("هل تم إرسال تقرير أسبوعي لمالك المحفظة؟", 4),
    ("هل تم التنبيه بثلاث أسابيع بحلول قرب تسليم المخرجات؟", 1),
    ("هل تم متابعة تسليم المخرجات؟", 1),
    ("هل تم تحديث الخطة المالية الرئيسية؟", 2),
    ("هل تم تصدير شهادات الإنجاز للمالية؟", 2),
    ("هل تم متابعة وتحديث المشاريع التشغيلية؟", 1),
    ("هل تم رفع جميع الوثائق في نظام ميسر؟", 1),
    ("هل يوجد مبادرات وإنجازات تم العمل عليها خلال هذا الأسبوع؟", 5),
    ("هل تم الالتزام بمصفوفة الصلاحيات التشغيلية والحوكمة واتباعها؟", 1),
]

PERFORMANCE_RATING = [
    ("0 - 70", 1, "Unsatisfactory", "غير مرضي: لم يحقق الأهداف الرئيسية أو معايير الأداء. يتطلب تحسين فوري."),
    ("71 - 85", 2, "Needs Improvement", "يحتاج تحسين: حقق بعض الأهداف لكن يوجد فجوات تتطلب تطوير إضافي."),
    ("86 - 100", 3, "Meets Expectations", "يلبي التوقعات: حقق جميع الأهداف المحددة وقدم نتائج متسقة مع الأداء المتوقع."),
    ("101 - 119", 4, "Exceeds Expectations", "يتجاوز التوقعات: حقق جميع الأهداف وقدم مبادرة ذات أثر قابل للقياس على مستوى القطاع."),
    ("120 - 130", 5, "Outstanding", "متميز: حقق نتائج استثنائية ذات أثر إيجابي على مستوى المركز. أظهر ابتكاراً وقيادة تتجاوز نطاق العمل."),
]

def style_cell(ws, row, col, value=None, font=None, fill=None, alignment=None, border=None, number_format=None, merge_end_col=None, merge_end_row=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if merge_end_col or merge_end_row:
        end_col = merge_end_col or col
        end_row = merge_end_row or row
        ws.merge_cells(start_row=row, start_column=col, end_row=end_row, end_column=end_col)
    return cell

def style_range(ws, start_row, start_col, end_row, end_col, font=None, fill=None, alignment=None, border=None):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border

def add_logo(ws, cell="A1"):
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH)
        img.width = 220
        img.height = 75
        ws.add_image(img, cell)

wb = Workbook()

# ============================================================
# SHEET 1: لوحة التحكم (Dashboard)
# ============================================================
ws_dash = wb.active
ws_dash.title = "لوحة التحكم"
ws_dash.sheet_view.rightToLeft = True

ws_dash.column_dimensions['A'].width = 5
for c in range(2, 10):
    ws_dash.column_dimensions[get_column_letter(c)].width = 22

add_logo(ws_dash, "H1")
ws_dash.row_dimensions[1].height = 30
ws_dash.row_dimensions[2].height = 30
ws_dash.row_dimensions[3].height = 30

style_cell(ws_dash, 1, 2, "مركز التأمين الصحي الوطني", Font(name='Calibri', bold=True, size=20, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=7)
style_cell(ws_dash, 2, 2, "Center for National Health Insurance", Font(name='Calibri', bold=True, size=14, color=TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=7)
style_cell(ws_dash, 3, 2, "لوحة التحكم - إدارة مشاريع PMO", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=9)
style_range(ws_dash, 3, 2, 3, 9, border=thin_border)

r = 5
style_cell(ws_dash, r, 2, "ملخص الأداء العام", subtitle_font, PatternFill('solid', fgColor=LIGHT_TEAL), center_align, merge_end_col=9, border=thin_border)
style_range(ws_dash, r, 2, r, 9, border=thin_border)

r = 6
kpi_labels = ["إجمالي الموظفين", "إجمالي المشاريع", "إجمالي المحافظ", "نسبة الإنجاز"]
kpi_values = [5, sum(len(e["projects"]) for e in EMPLOYEES), len(set(p for e in EMPLOYEES for p in e["portfolios"])), "0%"]
kpi_colors = [BLUE, GREEN, ORANGE, TEAL]

for i, (label, val, color) in enumerate(zip(kpi_labels, kpi_values, kpi_colors)):
    col = 2 + i * 2
    style_cell(ws_dash, r, col, label, Font(name='Calibri', bold=True, size=11, color=WHITE), PatternFill('solid', fgColor=color), center_align, thin_border, merge_end_col=col+1)
    style_cell(ws_dash, r, col+1, border=thin_border)
    style_cell(ws_dash, r+1, col, val, Font(name='Calibri', bold=True, size=20, color=color), PatternFill('solid', fgColor=LIGHT_GRAY), center_align, thin_border, merge_end_col=col+1)
    style_cell(ws_dash, r+1, col+1, border=thin_border)

r = 9
style_cell(ws_dash, r, 2, "بيانات الموظفين والمشاريع", subtitle_font, PatternFill('solid', fgColor=LIGHT_TEAL), center_align, merge_end_col=9, border=thin_border)
style_range(ws_dash, r, 2, r, 9, border=thin_border)

r = 10
emp_headers = ["#", "اسم الموظف", "المحافظ", "عدد المشاريع", "عدد المهام المنجزة", "نسبة الإنجاز", "الحالة", "التقييم"]
for i, h in enumerate(emp_headers):
    style_cell(ws_dash, r, 2+i, h, header_font, header_fill, center_align, thin_border)

for idx, emp in enumerate(EMPLOYEES):
    r = 11 + idx
    ws_dash.row_dimensions[r].height = 35
    style_cell(ws_dash, r, 2, idx+1, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_dash, r, 3, emp["name"], Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=center_align, border=thin_border)
    style_cell(ws_dash, r, 4, " / ".join(emp["portfolios"]), data_font, alignment=wrap_align, border=thin_border)
    style_cell(ws_dash, r, 5, len(emp["projects"]), Font(name='Calibri', bold=True, size=14, color=BLUE), alignment=center_align, border=thin_border)

    tasks_sheet_name = "المهام الأسبوعية"
    task_start_row = 4
    emp_task_offset = idx * (len(WEEKLY_TASKS) + 3)
    style_cell(ws_dash, r, 6, 0, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_dash, r, 7, "0%", data_font, alignment=center_align, border=thin_border)
    style_cell(ws_dash, r, 8, "جاري العمل", Font(name='Calibri', bold=True, size=11, color=ORANGE), alignment=center_align, border=thin_border)
    style_cell(ws_dash, r, 9, "-", data_font, alignment=center_align, border=thin_border)

r = 17
style_cell(ws_dash, r, 2, "توزيع المشاريع حسب الموظف", subtitle_font, PatternFill('solid', fgColor=LIGHT_TEAL), center_align, merge_end_col=9, border=thin_border)
style_range(ws_dash, r, 2, r, 9, border=thin_border)

chart_data_start = r + 1
for idx, emp in enumerate(EMPLOYEES):
    style_cell(ws_dash, chart_data_start + idx, 2, emp["name"], data_font, border=thin_border, alignment=center_align)
    style_cell(ws_dash, chart_data_start + idx, 3, len(emp["projects"]), data_font, border=thin_border, alignment=center_align)

chart = BarChart()
chart.type = "col"
chart.title = "عدد المشاريع لكل موظف"
chart.y_axis.title = "عدد المشاريع"
chart.style = 10
chart.width = 25
chart.height = 15
data_ref = Reference(ws_dash, min_col=3, min_row=chart_data_start, max_row=chart_data_start + 4)
cats_ref = Reference(ws_dash, min_col=2, min_row=chart_data_start, max_row=chart_data_start + 4)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = TEAL
chart.legend = None
ws_dash.add_chart(chart, "E18")

r = 35
style_cell(ws_dash, r, 2, "تعريف تقييم الأداء", subtitle_font, PatternFill('solid', fgColor=LIGHT_TEAL), center_align, merge_end_col=9, border=thin_border)
style_range(ws_dash, r, 2, r, 9, border=thin_border)

r = 36
perf_headers = ["مقياس الإنجاز", "الدرجة", "التصنيف", "التعريف"]
col_spans = [(2,3), (4,4), (5,5), (6,9)]
for (label, (sc, ec)) in zip(perf_headers, col_spans):
    style_cell(ws_dash, r, sc, label, header_font, header_fill, center_align, thin_border, merge_end_col=ec)
    for c in range(sc, ec+1):
        ws_dash.cell(row=r, column=c).border = thin_border

for idx, (scale, score, eng, ar) in enumerate(PERFORMANCE_RATING):
    row = 37 + idx
    ws_dash.row_dimensions[row].height = 40
    style_cell(ws_dash, row, 2, scale, data_font, alignment=center_align, border=thin_border, merge_end_col=3)
    style_cell(ws_dash, row, 3, border=thin_border)
    style_cell(ws_dash, row, 4, score, Font(name='Calibri', bold=True, size=14, color=DARK_TEAL), alignment=center_align, border=thin_border)
    style_cell(ws_dash, row, 5, eng, Font(name='Calibri', bold=True, size=10, color=DARK_GRAY), alignment=center_align, border=thin_border)
    style_cell(ws_dash, row, 6, ar, data_font, alignment=wrap_align, border=thin_border, merge_end_col=9)
    for c in range(6, 10):
        ws_dash.cell(row=row, column=c).border = thin_border

ws_dash.sheet_properties.tabColor = DARK_TEAL

print("Sheet 1: لوحة التحكم - Done")

# ============================================================
# SHEET 2: الموظفون (Employees)
# ============================================================
ws_emp = wb.create_sheet("الموظفون")
ws_emp.sheet_view.rightToLeft = True

ws_emp.column_dimensions['A'].width = 5
ws_emp.column_dimensions['B'].width = 8
ws_emp.column_dimensions['C'].width = 22
ws_emp.column_dimensions['D'].width = 45
ws_emp.column_dimensions['E'].width = 18
ws_emp.column_dimensions['F'].width = 18
ws_emp.column_dimensions['G'].width = 20
ws_emp.column_dimensions['H'].width = 20
ws_emp.column_dimensions['I'].width = 20
ws_emp.column_dimensions['J'].width = 20

add_logo(ws_emp, "I1")
ws_emp.row_dimensions[1].height = 30
ws_emp.row_dimensions[2].height = 30

style_cell(ws_emp, 1, 2, "مركز التأمين الصحي الوطني", Font(name='Calibri', bold=True, size=16, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=8)
style_cell(ws_emp, 2, 2, "إدارة الموظفين والمشاريع", Font(name='Calibri', bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=10)
style_range(ws_emp, 2, 2, 2, 10, border=thin_border)

r = 4
emp_detail_headers = ["#", "اسم الموظف", "المحافظ", "عدد المشاريع", "الإجازات السنوية المتبقية", "الإجازات المرضية المتبقية", "أيام العمل عن بعد المتبقية", "إجازات أخرى المتبقية", "الحالة"]
for i, h in enumerate(emp_detail_headers):
    style_cell(ws_emp, r, 2+i, h, header_font, header_fill, center_align, thin_border)

for idx, emp in enumerate(EMPLOYEES):
    row = 5 + idx
    ws_emp.row_dimensions[row].height = 40
    style_cell(ws_emp, row, 2, idx+1, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 3, emp["name"], Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 4, " / ".join(emp["portfolios"]), data_font, alignment=wrap_align, border=thin_border)
    style_cell(ws_emp, row, 5, len(emp["projects"]), Font(name='Calibri', bold=True, size=14, color=BLUE), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 6, 25, Font(name='Calibri', size=12, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 7, 60, Font(name='Calibri', size=12, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 8, 12, Font(name='Calibri', size=12, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 9, 0, Font(name='Calibri', size=12, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_emp, row, 10, "نشط", Font(name='Calibri', bold=True, size=11, color=GREEN), alignment=center_align, border=thin_border)

r = 5 + len(EMPLOYEES) + 2
style_cell(ws_emp, r, 2, "لإضافة موظف جديد:", Font(name='Calibri', bold=True, size=12, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=10)
style_cell(ws_emp, r+1, 2, "أضف بيانات الموظف الجديد في الصف التالي للموظفين الحاليين مباشرة", Font(name='Calibri', size=11, color=ORANGE), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=10)

ws_emp.sheet_properties.tabColor = "1565C0"
print("Sheet 2: الموظفون - Done")

# ============================================================
# SHEET 3: المهام الأسبوعية (Weekly Tasks)
# ============================================================
ws_tasks = wb.create_sheet("المهام الأسبوعية")
ws_tasks.sheet_view.rightToLeft = True

ws_tasks.column_dimensions['A'].width = 5
ws_tasks.column_dimensions['B'].width = 8
ws_tasks.column_dimensions['C'].width = 50
ws_tasks.column_dimensions['D'].width = 22
ws_tasks.column_dimensions['E'].width = 30
ws_tasks.column_dimensions['F'].width = 12
ws_tasks.column_dimensions['G'].width = 18

add_logo(ws_tasks, "F1")
style_cell(ws_tasks, 1, 2, "المهام الأسبوعية لكل مشروع", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=7)
style_range(ws_tasks, 1, 2, 1, 7, border=thin_border)

status_options = '"نعم,لا,لا ينطبق,جاري العمل عليه"'
dv_status = DataValidation(type="list", formula1=status_options, allow_blank=True)
dv_status.error = "يرجى اختيار قيمة من القائمة"
dv_status.errorTitle = "قيمة غير صحيحة"
dv_status.prompt = "اختر حالة المهمة"
dv_status.promptTitle = "حالة المهمة"
ws_tasks.add_data_validation(dv_status)

current_row = 3

for emp_idx, emp in enumerate(EMPLOYEES):
    style_cell(ws_tasks, current_row, 2, f"الموظف: {emp['name']}", Font(name='Calibri', bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, thin_border, merge_end_col=7)
    style_range(ws_tasks, current_row, 2, current_row, 7, border=thin_border)
    current_row += 1

    for proj_idx, (proj_name, proj_portfolio) in enumerate(emp["projects"]):
        style_cell(ws_tasks, current_row, 2, f"المشروع: {proj_name}", Font(name='Calibri', bold=True, size=11, color=WHITE), PatternFill('solid', fgColor=TEAL), Alignment(horizontal='right', vertical='center', wrap_text=True), thin_border, merge_end_col=5)
        style_range(ws_tasks, current_row, 2, current_row, 5, border=thin_border)
        style_cell(ws_tasks, current_row, 6, f"المحفظة: {proj_portfolio}", Font(name='Calibri', bold=True, size=10, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, thin_border, merge_end_col=7)
        style_range(ws_tasks, current_row, 6, current_row, 7, border=thin_border)
        current_row += 1

        task_headers = ["#", "المهمة", "الحالة", "الملاحظات", "ارتباط بالهدف", "حالة المشروع"]
        for i, h in enumerate(task_headers):
            style_cell(ws_tasks, current_row, 2+i, h, subheader_font, subheader_fill, center_align, thin_border)
        current_row += 1

        task_start = current_row
        for task_idx, (task_text, goal_link) in enumerate(WEEKLY_TASKS):
            ws_tasks.row_dimensions[current_row].height = 30
            style_cell(ws_tasks, current_row, 2, task_idx+1, data_font, alignment=center_align, border=thin_border)
            style_cell(ws_tasks, current_row, 3, task_text, data_font, alignment=wrap_align, border=thin_border)

            status_cell = ws_tasks.cell(row=current_row, column=4)
            status_cell.font = data_font
            status_cell.alignment = center_align
            status_cell.border = thin_border
            dv_status.add(status_cell)

            style_cell(ws_tasks, current_row, 5, "", data_font, alignment=wrap_align, border=thin_border)
            style_cell(ws_tasks, current_row, 6, f"الهدف {goal_link}", Font(name='Calibri', size=10, color=TEAL), alignment=center_align, border=thin_border)

            if task_idx == 0:
                task_end = task_start + len(WEEKLY_TASKS) - 1
                formula = f'=IF(COUNTIF(D{task_start}:D{task_end},"نعم")=COUNTA(D{task_start}:D{task_end}),"مكتمل ومغلق",IF(COUNTA(D{task_start}:D{task_end})=0,"لم يبدأ","جاري العمل عليه"))'
                style_cell(ws_tasks, current_row, 7, formula, Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=center_align, border=thin_border, merge_end_row=current_row + len(WEEKLY_TASKS) - 1)
            ws_tasks.cell(row=current_row, column=7).border = thin_border
            current_row += 1

        current_row += 1

    current_row += 1

green_fill = PatternFill('solid', fgColor="C8E6C9")
red_fill = PatternFill('solid', fgColor="FFCDD2")
orange_fill = PatternFill('solid', fgColor="FFE0B2")

ws_tasks.conditional_formatting.add(f"D3:D{current_row}", CellIsRule(operator='equal', formula=['"نعم"'], fill=green_fill))
ws_tasks.conditional_formatting.add(f"D3:D{current_row}", CellIsRule(operator='equal', formula=['"لا"'], fill=red_fill))
ws_tasks.conditional_formatting.add(f"D3:D{current_row}", CellIsRule(operator='equal', formula=['"جاري العمل عليه"'], fill=orange_fill))
ws_tasks.conditional_formatting.add(f"G3:G{current_row}", CellIsRule(operator='equal', formula=['"مكتمل ومغلق"'], fill=green_fill))
ws_tasks.conditional_formatting.add(f"G3:G{current_row}", CellIsRule(operator='equal', formula=['"جاري العمل عليه"'], fill=orange_fill))

ws_tasks.sheet_properties.tabColor = "43A047"
print("Sheet 3: المهام الأسبوعية - Done")

# ============================================================
# SHEET 4: الحضور والغياب (Attendance)
# ============================================================
ws_att = wb.create_sheet("الحضور والغياب")
ws_att.sheet_view.rightToLeft = True

ws_att.column_dimensions['A'].width = 5
ws_att.column_dimensions['B'].width = 22

add_logo(ws_att, "A1")
ws_att.row_dimensions[1].height = 30
ws_att.row_dimensions[2].height = 30

style_cell(ws_att, 2, 2, "سجل الحضور والغياب - 2026", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=28)
style_range(ws_att, 2, 2, 2, 28, border=thin_border)

r = 4
style_cell(ws_att, r, 2, "دليل الرموز", Font(name='Calibri', bold=True, size=13, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=10)

r = 5
icons_data = [
    ("\u2714", "حاضر", GREEN),
    ("\u2718", "غائب", RED),
    ("\U0001F912", "إجازة مرضية", ORANGE),
    ("\U0001F334", "إجازة سنوية", "2E7D32"),
    ("\U0001F3E0", "عمل عن بعد", BLUE),
    ("\u23F1", "إجازة أخرى", "7B1FA2"),
]

for i, (icon, label, color) in enumerate(icons_data):
    col = 2 + i * 2
    ws_att.column_dimensions[get_column_letter(col)].width = 6
    ws_att.column_dimensions[get_column_letter(col+1)].width = 14
    style_cell(ws_att, r, col, icon, Font(name='Segoe UI Emoji', size=16), alignment=center_align)
    style_cell(ws_att, r, col+1, label, Font(name='Calibri', bold=True, size=10, color=color), alignment=Alignment(horizontal='right', vertical='center'))

r = 7
style_cell(ws_att, r, 2, "رصيد الإجازات", Font(name='Calibri', bold=True, size=13, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, merge_end_col=10)
style_range(ws_att, r, 2, r, 10, border=thin_border)

r = 8
balance_headers = ["الموظف", "سنوية (25)", "المستخدمة", "المتبقية", "مرضية (60)", "المستخدمة", "المتبقية", "عن بعد (12)", "المستخدمة"]
for i, h in enumerate(balance_headers):
    style_cell(ws_att, r, 2+i, h, Font(name='Calibri', bold=True, size=10, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, thin_border)
    ws_att.column_dimensions[get_column_letter(2+i)].width = 16

for idx, emp in enumerate(EMPLOYEES):
    row = 9 + idx
    style_cell(ws_att, row, 2, emp["name"], Font(name='Calibri', bold=True, size=10, color=DARK_TEAL), alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 3, 25, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 4, 0, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 5, f"=C{row}-D{row}", Font(name='Calibri', bold=True, size=11, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 6, 60, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 7, 0, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 8, f"=F{row}-G{row}", Font(name='Calibri', bold=True, size=11, color=GREEN), alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 9, 12, data_font, alignment=center_align, border=thin_border)
    style_cell(ws_att, row, 10, 0, data_font, alignment=center_align, border=thin_border)

att_dv = DataValidation(type="list", formula1='"\u2714,\u2718,\U0001F912,\U0001F334,\U0001F3E0,\u23F1"', allow_blank=True)
att_dv.prompt = "اختر حالة الحضور"
att_dv.promptTitle = "الحضور"
ws_att.add_data_validation(att_dv)

import datetime
months_ar = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

att_row = 16
for month_idx in range(12):
    month_num = month_idx + 1
    style_cell(ws_att, att_row, 2, f"شهر {months_ar[month_idx]} 2026", Font(name='Calibri', bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=28)
    style_range(ws_att, att_row, 2, att_row, 28, border=thin_border)
    att_row += 1

    day_names = ["الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس"]
    working_days = []
    try:
        import calendar
        cal = calendar.monthcalendar(2026, month_num)
        for week in cal:
            for day_idx in [6, 0, 1, 2, 3]:
                if week[day_idx] != 0:
                    d = datetime.date(2026, month_num, week[day_idx])
                    if d.weekday() in [6, 0, 1, 2, 3]:
                        if week[day_idx] not in [wd[0] for wd in working_days]:
                            working_days.append((week[day_idx], day_names[[6,0,1,2,3].index(d.weekday())]))
    except:
        for day in range(1, 32):
            try:
                d = datetime.date(2026, month_num, day)
                if d.weekday() in [6, 0, 1, 2, 3]:
                    working_days.append((day, day_names[[6,0,1,2,3].index(d.weekday())]))
            except ValueError:
                break

    working_days.sort(key=lambda x: x[0])

    style_cell(ws_att, att_row, 2, "الموظف", Font(name='Calibri', bold=True, size=10, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, thin_border)

    for d_idx, (day_num, day_name) in enumerate(working_days):
        col = 3 + d_idx
        ws_att.column_dimensions[get_column_letter(col)].width = 6
        style_cell(ws_att, att_row, col, f"{day_num}\n{day_name}", Font(name='Calibri', bold=True, size=8, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, thin_border)
    att_row += 1

    for emp_idx, emp in enumerate(EMPLOYEES):
        style_cell(ws_att, att_row, 2, emp["name"], Font(name='Calibri', bold=True, size=9, color=DARK_TEAL), alignment=center_align, border=thin_border)
        for d_idx in range(len(working_days)):
            col = 3 + d_idx
            cell = ws_att.cell(row=att_row, column=col)
            cell.font = Font(name='Segoe UI Emoji', size=14)
            cell.alignment = center_align
            cell.border = thin_border
            att_dv.add(cell)
        att_row += 1

    att_row += 1

ws_att.sheet_properties.tabColor = "FB8C00"
print("Sheet 4: الحضور والغياب - Done")

# ============================================================
# SHEET 5: الأهداف السنوية (Annual Goals) - Protected
# ============================================================
ws_goals = wb.create_sheet("الأهداف السنوية")
ws_goals.sheet_view.rightToLeft = True

ws_goals.column_dimensions['A'].width = 5
ws_goals.column_dimensions['B'].width = 8
ws_goals.column_dimensions['C'].width = 22
ws_goals.column_dimensions['D'].width = 80
ws_goals.column_dimensions['E'].width = 15

add_logo(ws_goals, "D1")

style_cell(ws_goals, 1, 2, "الأهداف السنوية للموظفين - سري", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor="C62828"), center_align, merge_end_col=5)
style_range(ws_goals, 1, 2, 1, 5, border=thin_border)
style_cell(ws_goals, 2, 2, "هذه الصفحة محمية بكلمة مرور - للمدير فقط", Font(name='Calibri', bold=True, size=12, color="C62828"), alignment=Alignment(horizontal='center', vertical='center'), merge_end_col=5)

r = 4
goal_headers = ["#", "الموظف", "الهدف السنوي", "الوزن"]
for i, h in enumerate(goal_headers):
    style_cell(ws_goals, r, 2+i, h, header_font, PatternFill('solid', fgColor="C62828"), center_align, thin_border)

r = 5
for emp in EMPLOYEES:
    emp_start = r
    for g_idx, goal in enumerate(emp["goals"]):
        ws_goals.row_dimensions[r].height = 50
        style_cell(ws_goals, r, 2, g_idx + 1, data_font, alignment=center_align, border=thin_border)
        if g_idx == 0:
            style_cell(ws_goals, r, 3, emp["name"], Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=center_align, border=thin_border, merge_end_row=r+4)
        ws_goals.cell(row=r, column=3).border = thin_border
        style_cell(ws_goals, r, 4, goal, data_font, alignment=wrap_align, border=thin_border)
        style_cell(ws_goals, r, 5, "20%", data_font, alignment=center_align, border=thin_border)
        r += 1
    r += 1

ws_goals.protection.sheet = True
ws_goals.protection.password = "admin123"
ws_goals.protection.formatCells = False
ws_goals.protection.formatColumns = False
ws_goals.protection.formatRows = False

ws_goals.sheet_properties.tabColor = "C62828"
print("Sheet 5: الأهداف السنوية - Done")

# ============================================================
# SHEET 6: تقرير الطباعة (Print Report)
# ============================================================
ws_print = wb.create_sheet("تقرير الطباعة")
ws_print.sheet_view.rightToLeft = True

ws_print.column_dimensions['A'].width = 5
ws_print.column_dimensions['B'].width = 8
ws_print.column_dimensions['C'].width = 25
ws_print.column_dimensions['D'].width = 45
ws_print.column_dimensions['E'].width = 20
ws_print.column_dimensions['F'].width = 20
ws_print.column_dimensions['G'].width = 20

add_logo(ws_print, "F1")

style_cell(ws_print, 1, 2, "مركز التأمين الصحي الوطني", Font(name='Calibri', bold=True, size=16, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), merge_end_col=5)
style_cell(ws_print, 2, 2, "تقرير أداء الموظف", Font(name='Calibri', bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=7)
style_range(ws_print, 2, 2, 2, 7, border=thin_border)

r = 4
style_cell(ws_print, r, 2, "اختر الموظف:", Font(name='Calibri', bold=True, size=12, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'))
emp_names_list = ",".join([e["name"] for e in EMPLOYEES])
dv_emp = DataValidation(type="list", formula1=f'"{emp_names_list}"', allow_blank=True)
dv_emp.prompt = "اختر اسم الموظف"
ws_print.add_data_validation(dv_emp)
emp_select_cell = ws_print.cell(row=r, column=3)
emp_select_cell.font = Font(name='Calibri', bold=True, size=12, color=BLUE)
emp_select_cell.alignment = center_align
emp_select_cell.border = Border(bottom=Side(style='thick', color=DARK_TEAL))
dv_emp.add(emp_select_cell)

print_row = 6
for emp_idx, emp in enumerate(EMPLOYEES):
    start_row = print_row
    style_cell(ws_print, print_row, 2, f"تقرير الموظف: {emp['name']}", Font(name='Calibri', bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=7)
    style_range(ws_print, print_row, 2, print_row, 7, border=thin_border)
    print_row += 1

    style_cell(ws_print, print_row, 2, "البيانات الأساسية", Font(name='Calibri', bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, merge_end_col=7)
    style_range(ws_print, print_row, 2, print_row, 7, border=thin_border)
    print_row += 1

    info_items = [
        ("اسم الموظف", emp["name"]),
        ("المحافظ", " / ".join(emp["portfolios"])),
        ("عدد المشاريع", len(emp["projects"])),
    ]
    for label, val in info_items:
        style_cell(ws_print, print_row, 2, label, Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), border=thin_border, merge_end_col=3)
        ws_print.cell(row=print_row, column=3).border = thin_border
        style_cell(ws_print, print_row, 4, val, data_font, alignment=wrap_align, border=thin_border, merge_end_col=7)
        for c in range(4, 8):
            ws_print.cell(row=print_row, column=c).border = thin_border
        print_row += 1

    print_row += 1
    style_cell(ws_print, print_row, 2, "المشاريع", Font(name='Calibri', bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=TEAL), center_align, merge_end_col=7)
    style_range(ws_print, print_row, 2, print_row, 7, border=thin_border)
    print_row += 1

    proj_headers = ["#", "اسم المشروع", "", "المحفظة", "الحالة", ""]
    for i, h in enumerate(proj_headers):
        style_cell(ws_print, print_row, 2+i, h, subheader_font, subheader_fill, center_align, thin_border)
    print_row += 1

    for p_idx, (p_name, p_port) in enumerate(emp["projects"]):
        style_cell(ws_print, print_row, 2, p_idx+1, data_font, alignment=center_align, border=thin_border)
        style_cell(ws_print, print_row, 3, p_name, data_font, alignment=wrap_align, border=thin_border, merge_end_col=4)
        ws_print.cell(row=print_row, column=4).border = thin_border
        style_cell(ws_print, print_row, 5, p_port, data_font, alignment=center_align, border=thin_border)
        style_cell(ws_print, print_row, 6, "جاري العمل", Font(name='Calibri', size=10, color=ORANGE), alignment=center_align, border=thin_border, merge_end_col=7)
        ws_print.cell(row=print_row, column=7).border = thin_border
        ws_print.row_dimensions[print_row].height = 30
        print_row += 1

    print_row += 1
    style_cell(ws_print, print_row, 2, "ملاحظة: الأهداف السنوية غير معروضة في هذا التقرير", Font(name='Calibri', italic=True, size=10, color=MEDIUM_GRAY), alignment=Alignment(horizontal='center', vertical='center'), merge_end_col=7)
    print_row += 2

ws_print.print_area = f"A1:G{print_row}"
ws_print.page_setup.orientation = 'portrait'
ws_print.page_setup.fitToWidth = 1
ws_print.page_setup.fitToHeight = 0
ws_print.page_setup.paperSize = ws_print.PAPERSIZE_A4
ws_print.sheet_properties.tabColor = "7B1FA2"
print("Sheet 6: تقرير الطباعة - Done")

# ============================================================
# SHEET 7: المشاريع (Projects)
# ============================================================
ws_proj = wb.create_sheet("المشاريع")
ws_proj.sheet_view.rightToLeft = True

ws_proj.column_dimensions['A'].width = 5
ws_proj.column_dimensions['B'].width = 8
ws_proj.column_dimensions['C'].width = 60
ws_proj.column_dimensions['D'].width = 30
ws_proj.column_dimensions['E'].width = 22
ws_proj.column_dimensions['F'].width = 22

add_logo(ws_proj, "E1")

style_cell(ws_proj, 1, 2, "سجل المشاريع", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=6)
style_range(ws_proj, 1, 2, 1, 6, border=thin_border)

r = 3
proj_headers = ["#", "اسم المشروع", "المحفظة", "المسؤول", "الحالة"]
for i, h in enumerate(proj_headers):
    style_cell(ws_proj, r, 2+i, h, header_font, header_fill, center_align, thin_border)

r = 4
proj_num = 1
proj_status_dv = DataValidation(type="list", formula1='"مكتمل ومغلق,جاري العمل عليه,متوقف,لم يبدأ"', allow_blank=True)
ws_proj.add_data_validation(proj_status_dv)

for emp in EMPLOYEES:
    for p_name, p_port in emp["projects"]:
        ws_proj.row_dimensions[r].height = 30
        style_cell(ws_proj, r, 2, proj_num, data_font, alignment=center_align, border=thin_border)
        style_cell(ws_proj, r, 3, p_name, data_font, alignment=wrap_align, border=thin_border)
        style_cell(ws_proj, r, 4, p_port, data_font, alignment=center_align, border=thin_border)
        style_cell(ws_proj, r, 5, emp["name"], Font(name='Calibri', bold=True, size=10, color=DARK_TEAL), alignment=center_align, border=thin_border)

        status_cell = ws_proj.cell(row=r, column=6)
        status_cell.value = "جاري العمل عليه"
        status_cell.font = Font(name='Calibri', bold=True, size=10, color=ORANGE)
        status_cell.alignment = center_align
        status_cell.border = thin_border
        proj_status_dv.add(status_cell)

        proj_num += 1
        r += 1

ws_proj.conditional_formatting.add(f"F4:F{r}", CellIsRule(operator='equal', formula=['"مكتمل ومغلق"'], fill=green_fill))
ws_proj.conditional_formatting.add(f"F4:F{r}", CellIsRule(operator='equal', formula=['"جاري العمل عليه"'], fill=orange_fill))
ws_proj.conditional_formatting.add(f"F4:F{r}", CellIsRule(operator='equal', formula=['"متوقف"'], fill=red_fill))

r += 2
style_cell(ws_proj, r, 2, "ملخص المشاريع", subtitle_font, PatternFill('solid', fgColor=LIGHT_TEAL), center_align, merge_end_col=6)
style_range(ws_proj, r, 2, r, 6, border=thin_border)
r += 1

summary_items = [
    ("إجمالي المشاريع", f"=COUNTA(C4:C{r-3})"),
    ("مكتمل ومغلق", f'=COUNTIF(F4:F{r-3},"مكتمل ومغلق")'),
    ("جاري العمل عليه", f'=COUNTIF(F4:F{r-3},"جاري العمل عليه")'),
    ("متوقف", f'=COUNTIF(F4:F{r-3},"متوقف")'),
]
for label, formula in summary_items:
    style_cell(ws_proj, r, 2, label, Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), alignment=Alignment(horizontal='right', vertical='center'), border=thin_border, merge_end_col=4)
    for c in range(2, 5):
        ws_proj.cell(row=r, column=c).border = thin_border
    style_cell(ws_proj, r, 5, formula, Font(name='Calibri', bold=True, size=14, color=DARK_TEAL), alignment=center_align, border=thin_border, merge_end_col=6)
    ws_proj.cell(row=r, column=6).border = thin_border
    r += 1

ws_proj.sheet_properties.tabColor = "43A047"
print("Sheet 7: المشاريع - Done")

# ============================================================
# SHEET 8: تقييم الأداء (Performance Rating)
# ============================================================
ws_perf = wb.create_sheet("تقييم الأداء")
ws_perf.sheet_view.rightToLeft = True

ws_perf.column_dimensions['A'].width = 5
ws_perf.column_dimensions['B'].width = 22
ws_perf.column_dimensions['C'].width = 15
ws_perf.column_dimensions['D'].width = 25
ws_perf.column_dimensions['E'].width = 70

add_logo(ws_perf, "D1")

style_cell(ws_perf, 1, 2, "تعريف تقييم الأداء - Performance Rating Definition", Font(name='Calibri', bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_TEAL), center_align, merge_end_col=5)
style_range(ws_perf, 1, 2, 1, 5, border=thin_border)

r = 3
perf_detail_headers = ["مقياس الإنجاز\nAchievement Scale", "الدرجة\nRating Score", "التصنيف\nRating", "التعريف\nRating Definitions"]
for i, h in enumerate(perf_detail_headers):
    style_cell(ws_perf, r, 2+i, h, header_font, header_fill, center_align, thin_border)
    ws_perf.row_dimensions[r].height = 40

for idx, (scale, score, eng, ar) in enumerate(PERFORMANCE_RATING):
    row = 4 + idx
    ws_perf.row_dimensions[row].height = 55
    row_fill = PatternFill('solid', fgColor=LIGHT_TEAL) if idx % 2 == 0 else PatternFill('solid', fgColor=WHITE)

    style_cell(ws_perf, row, 2, scale, Font(name='Calibri', bold=True, size=12, color=DARK_GRAY), row_fill, center_align, thin_border)
    style_cell(ws_perf, row, 3, score, Font(name='Calibri', bold=True, size=16, color=DARK_TEAL), row_fill, center_align, thin_border)
    style_cell(ws_perf, row, 4, eng, Font(name='Calibri', bold=True, size=11, color=DARK_TEAL), row_fill, center_align, thin_border)
    style_cell(ws_perf, row, 5, ar, data_font, row_fill, wrap_align, thin_border)

ws_perf.sheet_properties.tabColor = "FF6F00"
print("Sheet 8: تقييم الأداء - Done")

# ============================================================
# Save workbook
# ============================================================
wb.save(OUTPUT_PATH)
print(f"\nDashboard saved to: {OUTPUT_PATH}")
print("Done!")
